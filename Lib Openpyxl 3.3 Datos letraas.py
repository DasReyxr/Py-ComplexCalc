from array import array
import string
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font 
import time

array_datos=["a","b","c","d","e","f","g","h","i"]
#Carga archivo con el nombre
book = load_workbook("ExcelTest3.xlsx")
#Cambiar valor
sheet = book["Hoja1"]

#Crea archivo
#book = Workbook()
# celda2 = input("Ingrese letraCelda: ")
# celda = input("Ingrese numero celda: ")
# valor = input("Ingrese valor: ")

# sheet[celda2+celda] =  valor

# #Acceder valor
# celdaobtenida = sheet[celda2+celda]
# print(celdaobtenida.value)

#Cambia tabla izq derecha
Inicial =input("Ingrese Inicial: ")
Final =input("Ingrese Final: ")
print(f"{list(Inicial)[0]}= {ord(list(Inicial)[0].lower())-96}")
numero_init=int(list(Inicial)[1])

print(f"{list(Final)[0]}= {ord(list(Final)[0].lower())-96}")
numero_fin=int(list(Final)[1]+list(Final)[2])
print(numero_fin)
rango = sheet[Inicial:Final]
longitud=len(rango)
print(f"Longitud: {longitud}")
#print(len(array_datos))
numero=0
for fila in range(numero_init,numero_fin+1):
    for col in range(ord(list(Inicial)[0].lower())-96,ord(list(Final)[0].lower())-96+1):
            num_letra=chr(col+96).upper() #Celda columna            
            if(len(array_datos)<=numero):
                sheet[num_letra+str(fila)] = ""    
            else :
                sheet[num_letra+str(fila)] = array_datos[numero]
            
            print(f"{sheet[num_letra+str(fila)].value}")
            numero=numero+1



book.save("ExcelTest3.xlsx")