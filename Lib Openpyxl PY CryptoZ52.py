from array import array
import string
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font 
import time
book = Workbook()
array_abc=[0,"a","b","c","d","e","f","g","h","i","j","k","l","m","n","o","p","q","r","s","t","u","v","w","x","y","z"]
array_cba=[]

array_cbc=array_abc[27:1:-1]
print(array_cbc)
#Carga archivo con el nombre
#book = load_workbook("ExcelTestABC.xlsx")
#Cambiar valor
sheet = book["Sheet"]
#sheet = book.active
#Cambia tabla izq derecha
Inicial = input("Ingrese Inicial: ")
Final =   input("Ingrese Final: ")

def arrayfunct(Inicial,Final,array_datos,array_save):
    separado_init= list(Inicial)
    letra_init=separado_init[0]
    letra_num=ord(letra_init.lower())-96
    print(f"{letra_init}= {letra_num}")
    numero_init=int(separado_init[1])

    separado_fin= list(Final)
    letra_fin=separado_fin[0]
    letra_num2=ord(letra_fin.lower())-96

    print(f"{letra_fin}= {letra_num2}")
    numero_fin=int(separado_fin[1])
    rango = sheet[Inicial:Final]
    longitud=len(rango)
    print(f"Longitud: {longitud}")
    print(len(array_datos))
    numero=0
    for fila in range(numero_init,numero_fin+1):
        for col in range(letra_num,letra_num2+1):
                num_letra=chr(col+96).upper() #Celda columna   
                   
                if(len(array_datos)-1<=numero):
                    sheet[num_letra+str(fila)] = ""    
                else :
                    sheet[num_letra+str(fila)] = array_datos[numero+1]
                    #Save
                    #array_save.insert(numero,sheet[num_letra+str(fila)].value)
                    #prnt
                    sheet[array_datos[col+5]+str(fila)] = array_save[numero-1]
                
                # print(f"{sheet[num_letra+str(fila)].value}")
                numero=numero+1


# # def excelfunct(Inicial,Final,array_datos):
#     separado_init= list(Inicial)
#     letra_init=separado_init[0]
#     letra_num=ord(letra_init.lower())-96
#     print(f"{letra_init}= {letra_num}")
#     numero_init=int(separado_init[1])

#     separado_fin= list(Final)
#     letra_fin=separado_fin[0]
#     letra_num2=ord(letra_fin.lower())-96
#     print(f"{letra_fin}= {letra_num2}")
#     numero_fin=int(separado_fin[1])
#     #rango = sheet[Inicial:Final]
#     print(f"Longitud: {len(array_datos)}")
#     numero=0
#     for fila in range(numero_init,numero_fin+1):
#         for col in range(letra_num,letra_num2+1):
#                 num_letra=chr(col+96).upper() #Celda columna            
#                 # if(len(array_datos)<=numero):
#                 #     sheet[num_letra+str(fila)] = ""    
#                 # else :
#                 array_datos.insert(numero,sheet[num_letra+str(fila)].value)
                
#                 print(f"{sheet[num_letra+str(fila)].value}")
#                 numero=numero+1

arrayfunct(Inicial,Final,array_abc,array_cbc)

print(array_cbc)
book.save("ExcelTestABC.xlsx")