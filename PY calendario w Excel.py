from array import ArrayType
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment

import time
FILE_NAME ="ExcelCal1.xlsx"
DOW_CELL_INIT = "C1"
DOW_CELL_END = "I1"
book = Workbook()
sheet = book.active
#book = load_workbook(FILE_NAME)
#sheet = book["Sheet"]

thin= Side(border_style="thin")
marco= Border(top=thin,right=thin,bottom=thin,left=thin)

sheet.column_dimensions["A"].width = 7
sheet.column_dimensions["B"].width = 7
INIT ="C2"
END ="I44"
num_init=1

black ="000000"
white = "F2F2F2"
gray = "808080" 
green =  "00B050"
cell_white=PatternFill(patternType="solid",
                fgColor=white)
cell_gray= PatternFill(patternType="solid",
                fgColor=gray)
cell_green= PatternFill(patternType="solid",
                fgColor=green)

for row in sheet["C1:I44"]:
    for cell in row:
        cell.border = marco



#clear sheet
for row in sheet['A1:Z100']:
  for cell in row:
    cell.value = None
    cell.fill = cell_white
    cell.font = Font(color=black)

dias =[0,31,28,31,30,31,30,31,31,30,31,30,31]
dow = [0,"Lun","Mar","Mie","Jue","Vie","Sab","Dom"]
turno =["Night","Night","Free","Free","Day","Day"]
array_mes =[0,"Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
dia = {
    "mes": 0
}
dia["dow"]=[]
dia["dianum"]=[]

def loop():
    mes_ingresado = int(time.strftime("%m"))
    dia_ingresado= int(time.strftime("%d"))
    print(f"Mes {mes_ingresado}) {array_mes[mes_ingresado]}")
    #mes_ingresado = int(input("Ingresa mes: "))
    print(f"Dia {dia_ingresado} inicia en: ")
    # for i in range(1,8):
    #     print(f"{i}) {dow[i]}")
    dow_ingresado = int(time.strftime("%w"))
    print(f"{dow_ingresado}) {dow[dow_ingresado]} \n")
    #dow_ingresado = int(input("Ingrese dow: "))
    for i in range(0,6):
        print(f"{i}) {turno[i]}")
    turno_ingresado = int(input("Ingrese turno: "))
    
    Inst_mes = Mes()
    Inst_mes.fun_mes(mes_ingresado,dow_ingresado,turno_ingresado)
    Inst_mes.prnt_mes()
    dia["mes"] = mes_ingresado

 #   print(f"cancion: {dia}")

    
    Inst_mes.prnt_dias(dow_ingresado,turno_ingresado,dia_ingresado)
    Inst_mes.prnt_dow(dow)
    #Inst_mes.set_colors(turno_ingresado)
    # for dowcont in dia["dow"]:
    #     print(dowcont)
    book.save(FILE_NAME)

class Mes:
    def fun_mes(self,mes,dow_init,turn_init):
        self.mes = mes
        self.dow_init = dow_init
        self.turn_init = turn_init
    def prnt_mes(self):
        print(f"Mes: {array_mes[self.mes]}")
    def prnt_dow(self,array_dow):
        self.array_dow=array_dow
        Primer_letra=ord(list(INIT)[0].lower())-96
        Ult_Letra =ord(list(END)[0].lower())-96
        for char in range(Primer_letra,Ult_Letra+1):
                num_letra=chr(char+96).upper() #Celda columna  
                sheet[f"{num_letra}1"] =dow[char-2]
                      
    def prnt_dias(self,dow_init,turn_init,num_init):
        vuelta = False
        self.dow_init=dow_init
        
        numero_init=int(list(INIT)[1])
        numero_fin=int(list(END)[1]+list(END)[2])

        letra_fin = ord(list(END)[0].lower())-96
        for fila in range(numero_init,numero_fin+1):
            if (vuelta==False):
                letra_init=ord(list(INIT)[0].lower())-96+dow_init-1
            else:
                letra_init = ord(list(INIT)[0].lower())-96
            for col in range(letra_init,letra_fin+1):
                num_letra=chr(col+96).upper() #Celda columna
                if(self.dow_init>7):
                    self.dow_init=1
                if(self.turn_init==6):
                    self.turn_init=0
                
                if(num_init>(dias[self.mes])):
                    cc1=f"A"+str(fila-3)
                    cc2=f"B"+str(fila)
                    sheet[cc1] = array_mes[self.mes]
                    sheet[cc1].border = marco
                    sheet.merge_cells(f"{cc1}:{cc2}")
                    sheet[cc1].alignment = Alignment(horizontal="center",vertical="center")
                    
                    # sheet.merge_cells(f"{cc2}:{cc4}")
                    # sheet.merge_cells(f"{cc1}:{cc3}")
                    if(self.mes==12):
                        self.mes=1
                        num_init=1    
                    else:
                        self.mes=self.mes+1
                        num_init=1
                
                    #size sheet
                
                sheet.column_dimensions[num_letra].width = 7
                sheet[f"{num_letra}"+str(fila)] = num_init
                # if (col==letra_fin):
                #     print("")
                #print(f"{num_init} {dow[self.dow_init]} {turno[self.turn_init]}",end =" ")
                if (self.turn_init==0):
                    sheet[f"{num_letra}"+str(fila)].fill = cell_gray
                    sheet[f"{num_letra}"+str(fila)].font = Font(color=white)
                    
                elif (self.turn_init==1):
                    
                    sheet[f"{num_letra}"+str(fila)].fill = cell_gray
                    sheet[f"{num_letra}"+str(fila)].font = Font(color=white)
                elif (self.turn_init==2):
                    sheet[f"{num_letra}"+str(fila)].fill = cell_green
                    sheet[f"{num_letra}"+str(fila)].font = Font(color=white)
                    
                elif (self.turn_init==3):
                    sheet[f"{num_letra}"+str(fila)].fill = cell_green
                    sheet[f"{num_letra}"+str(fila)].font = Font(color=white)
                elif (self.turn_init==4):
                    sheet[f"{num_letra}"+str(fila)].fill = cell_white
                    sheet[f"{num_letra}"+str(fila)].font = Font(color=black)
                elif (self.turn_init==5):
                    sheet[f"{num_letra}"+str(fila)].fill = cell_white
                    sheet[f"{num_letra}"+str(fila)].font = Font(color=black)
                
                self.dow_init=self.dow_init+1
                #dia["dianum"].append(self.turn_init)
                self.turn_init=self.turn_init+1
                num_init=num_init+1
            vuelta = True





loop()