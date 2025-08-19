from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font 
import time

#Crea archivo
#book = Workbook()
#Carga archivo con el nombre
book = load_workbook("ExcelTest3.xlsx")
sheet = book["Hoja1"]

#Array->Excel
rango = sheet["A2":"D5"]
array=[["a","b","c","d","e"]["1","2","3","4","5"]]
longitud=len(rango)
print(f"Longitud: {longitud}")

#Imprime tabla izq-der
# rango_fila = sheet["A2":"A5"]
# rango_columna = sheet["D2":"D5"]
# longitud_fila=len(rango_fila)
# longitud_columna=len(rango_columna)
rango = sheet["A2":"D5"]
longitud=len(rango)
print(f"Longitud: {longitud}")
for fila in range(0,longitud):
    for col in range(0,longitud):
        print(f"({fila},{col}) = {rango[fila][col].value}")

# #Imprime Fila
# rango = sheet["A2":"D5"]
# longitud=len(rango)
# for i in range(0,longitud):
#     print(rango[i][0].value)

# print("Imprime columna ")
# #Imprime columna
# rango = sheet["D2":"D5"]
# longitud=len(rango)
# for i in range(0,longitud):
#     print(rango[i][0].value)
# #Acceder valor



book.save("ExcelTest3.xlsx")