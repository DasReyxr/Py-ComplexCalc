
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from win32com import client
from tkinter import *#Import all of the clases
import os
import time
FILE_NAME ="ExcelCal2"
EXT = ".xlsx"
DOW_CELL_INIT = "C1"
DOW_CELL_END = "I1"
ROOT = Tk()
input_tk=StringVar()

book = Workbook()
sheet = book.active
#-------En caso de que ya lo haya generado comentar las dos lineas superiores
#---- Y descomentar las dos lineas inferiores
#book = load_workbook(FILE_NAME+EXT)
#heet = book["Sheet"]

thin= Side(border_style="thin")
marco= Border(top=thin,right=thin,bottom=thin,left=thin)

sheet.column_dimensions["A"].width = 4.5
sheet.column_dimensions["B"].width = 4.5
INIT ="C2"
END ="I44"


black ="000000"
white = "FFFFFF"
gray = "808080" 
green =  "00B050"
cell_white=PatternFill(patternType="solid",
                fgColor=white)
cell_gray= PatternFill(patternType="solid",
                fgColor=gray)
cell_green= PatternFill(patternType="solid",
                fgColor=green)

for row in sheet["C1:I44"]:
    for cell in row:
        cell.border = marco

#clear sheet
for row in sheet['A1:I44']:
  for cell in row:
    cell.value = None
    cell.fill = cell_white
    cell.font = Font(color=black)

dias =[0,31,28,31,30,31,30,31,31,30,31,30,31]
dow = [0,"Lun","Mar","Mie","Jue","Vie","Sab","Dom"]
turno =["Night","Night","Free","Free","Day","Day"]
array_mes =[0,"Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

#Tkinter
FRAME=Frame(ROOT,width=1200,height=600)
FRAME.pack(fill="both",expand="True")

col=0
row=0
mes_ingresado = int(time.strftime("%m"))
dia_ingresado= int(time.strftime("%d"))
dow_ingresado = int(time.strftime("%w"))
text_dia=f"Dia {dia_ingresado} inicia en: "
ET_LAB =Label(FRAME,text=text_dia) 
ET_LAB.grid(row=row,column=col,columnspan=3,sticky="w")

row=4
range_dow=f"{0}) {turno[0]}"
dow_0 =Label(FRAME,text=range_dow) 
dow_0.grid(row=row,column=col,columnspan=3,sticky="w")
row+=1
range_dow=f"{1}) {turno[1]}"
dow_1 =Label(FRAME,text=range_dow) 
dow_1.grid(row=row,column=col,columnspan=3,sticky="w")
row+=1
range_dow=f"{2}) {turno[2]}"
dow_2 =Label(FRAME,text=range_dow) 
dow_2.grid(row=row,column=col,columnspan=3,sticky="w")
row+=1
range_dow=f"{3}) {turno[3]}"
dow_3 =Label(FRAME,text=range_dow) 
dow_3.grid(row=row,column=col,columnspan=3,sticky="w")
row+=1
range_dow=f"{4}) {turno[4]}"
dow_4 =Label(FRAME,text=range_dow) 
dow_4.grid(row=row,column=col,columnspan=3,sticky="w")
row+=1
range_dow=f"{5}) {turno[5]}"
dow_5 =Label(FRAME,text=range_dow) 
dow_5.grid(row=row,column=col,columnspan=3,sticky="w")
row+=1




# for i in range(0,6):
#     range_dow=f"{i}) {turno[i]}"
#     ET_LAB2 =Label(FRAME,text=range_dow) 
#     ET_LAB2.grid(row=row,column=col,columnspan=3,sticky="w")
#     row+=1

#----Boton
col=0
row=1

ET_LAB3 =Label(FRAME,text="Turno")
ET_LAB3.grid(row=row,column=col)
col+=1
INPUT = Entry(FRAME,textvariable=input_tk)
INPUT.grid(row=row,column=col)

def butontkinter():
    loop(input_tk.get())

def loop(turn_tk):
    global FRAME;global row;global col;global LABEL2
    global mes_ingresado;global dow_ingresado;global dia_ingresado
    turno_ingresado=int(turn_tk)
    col=0;row=0
    text_dia2=f"Dia {dia_ingresado} inicia en: {dow[turno_ingresado+1]}"
    ET_LAB4 =Label(FRAME,text=text_dia2) 
    ET_LAB4.grid(row=row,column=col,columnspan=3,sticky="w")
    Inst_mes = Mes()
    Inst_mes.prnt_dias(mes_ingresado,dow_ingresado,turno_ingresado,dia_ingresado)
    Inst_mes.prnt_dow(dow)
    book.save(FILE_NAME+EXT)
    currentDir = os.getcwd()

    # # Open Microsoft Excel
    excel = client.Dispatch("Excel.Application")
    books = excel.Workbooks.Open(os.path.join(currentDir,FILE_NAME+EXT))    
    ws = books.Worksheets[0]
    ws.Visible = 1
    DESTINY=os.path.join(currentDir,FILE_NAME+".pdf")
    ws.ExportAsFixedFormat(0,DESTINY )
    books.Close()
    text_dest=f"PPTX to PDF {DESTINY}"
    row=3;col=0
    ET_LAB5 =Label(FRAME,text=text_dest) 
    ET_LAB5.grid(row=row,column=col,columnspan=3,sticky="w")
    dow_0.grid_forget()
    dow_1.grid_forget()
    dow_2.grid_forget()
    dow_3.grid_forget()
    dow_4.grid_forget()
    dow_5.grid_forget()
    
    
    

class Mes:
    def prnt_dow(self,array_dow):
        self.array_dow=array_dow
        Primer_letra=ord(list(INIT)[0].lower())-96
        Ult_Letra =ord(list(END)[0].lower())-96
        for char in range(Primer_letra,Ult_Letra+1):
                num_letra=chr(char+96).upper() #Celda columna  
                sheet[f"{num_letra}1"] =dow[char-2]
                      
    def prnt_dias(self,mes,dow_init,turn_init,num_init):
        vuelta = False
        self.mes = mes
        self.dow_init = dow_init  
        self.turn_init = turn_init    
        numero_init=int(list(INIT)[1])
        numero_fin=int(list(END)[1]+list(END)[2])
        letra_fin = ord(list(END)[0].lower())-96
        for fila in range(numero_init,numero_fin+1):
            if (vuelta==False):
                letra_init=ord(list(INIT)[0].lower())-96+dow_init-1
            else:
                letra_init = ord(list(INIT)[0].lower())-96
            for col in range(letra_init,letra_fin+1):
                num_letra=chr(col+96).upper() #Celda columna
                if(self.dow_init>7):
                    self.dow_init=1
                if(self.turn_init==6):
                    self.turn_init=0
                
                if(num_init>(dias[self.mes])):
                    cc1=f"A"+str(fila)
                    cc2=f"B"+str(fila)
                    sheet[cc1] = array_mes[self.mes]
                    sheet[cc1].border = marco
                    sheet[cc1].alignment = Alignment(horizontal="center")
                    sheet.merge_cells(f"{cc1}:{cc2}")
                    if(self.mes==12):
                        self.mes=1
                        num_init=1    
                    else:
                        self.mes=self.mes+1
                        num_init=1
                cell= f"{num_letra}"+str(fila)
                sheet.column_dimensions[num_letra].width = 7
                sheet[cell] = num_init
                if (self.turn_init==0):
                    sheet[cell].fill = cell_gray
                    sheet[cell].font = Font(color=white)           
                elif (self.turn_init==1):        
                    sheet[cell].fill = cell_gray
                    sheet[cell].font = Font(color=white)
                elif (self.turn_init==2):
                    sheet[cell].fill = cell_green
                    sheet[cell].font = Font(color=white)
                elif (self.turn_init==3):
                    sheet[cell].fill = cell_green
                    sheet[cell].font = Font(color=white)
                elif (self.turn_init==4):
                    sheet[cell].fill = cell_white
                    sheet[cell].font = Font(color=black)
                elif (self.turn_init==5):
                    sheet[cell].fill = cell_white
                    sheet[cell].font = Font(color=black)
                self.dow_init=self.dow_init+1
                self.turn_init=self.turn_init+1
                num_init=num_init+1
            vuelta = True

EPA=Button(FRAME,text="Send",command=butontkinter)
EPA.grid(row=row,column=3)

ROOT.mainloop() #Infinite loop 