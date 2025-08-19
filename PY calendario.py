from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font 
import time
# FILE_NAME ="ExcelCal1.xlsx"
# book = Workbook()
# sheet = book.active
#book = load_workbook(FILE_NAME)
#sheet = book["Sheet"]

gray = "808080" 
green =  "00B050"

#red= PatternFill(patternType="solid",
 #               fgColor="C64747")

dias =[0,31,28,31,30,31,30,31,31,30,31,30,31]
dow = ["Dom","Lun","Mar","Mie","Jue","Vie","Sab"]
turno =["Night","Night","Free","Free","Day","Day"]
array_mes =[0,"Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
dia = {
    "mes": 0
}
dia["dow"]=[]
dia["dianum"]=[]

def loop():
    mes_ingresado = int(input("Ingresa mes: "))
    print("Dia inicial: ")
    for i in range(0,7):
        print(f"{i}) {dow[i]}")
    dia_ingresado = int(input("Ingrese dow: "))
    for i in range(0,6):
        print(f"{i}) {turno[i]}")
    turno_ingresado = int(input("Ingrese turno: "))
    
    Inst_mes = Mes()
    Inst_mes.fun_mes(mes_ingresado,dia_ingresado,turno_ingresado)
    dia["mes"] = mes_ingresado

    print(f"cancion: {dia}")

    Inst_mes.prnt_mes()
    Inst_mes.prnt_dias()
    # for dowcont in dia["dow"]:
    #     print(dowcont)
    # book.save(FILE_NAME)

class Mes:
    def fun_mes(self,mes,dow_init,turn_init):
        self.mes = mes
        self.dow_init = dow_init
        self.turn_init = turn_init
    def prnt_mes(self):
        print(f"Mes: {array_mes[self.mes]}")
    def prnt_dias(self):
        for i in range(1,dias[self.mes]+1):
            if (self.dow_init==7):
                self.dow_init=0
            if (self.turn_init==6):
                self.turn_init=0
            if((i)%7==0):
                print("")
                
            print(f"{i} {dow[self.dow_init]} {turno[self.turn_init]}",end = ' ')
            dia["dow"].append(dow[self.dow_init]) #Agrega
            dia["dianum"].append(self.dow_init)
            self.turn_init=self.turn_init+1
            self.dow_init=self.dow_init+1
            i=i+1


loop()