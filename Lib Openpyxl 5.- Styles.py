from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill 
from openpyxl.styles import Font 
from openpyxl.styles import Border
from openpyxl.styles import Side
from win32com import client
from openpyxl.styles import borders
import time
import os



array_datos=[0,"a","A","b","B","c","d","H","I","j","m","M","p","S","U","w","W","x","X","y","Y","Z","%"]
array_datos2=[0,"a","b","c","d","e","f","g","h","i","j","k","l","m","n","o","p","q","r","s","t","u","v","w","x","y","z"]
FILE = "ExcelTest4a.xlsx"

FILE_SHEET2 = "Hoja21"
FILE_SHEET3 = "dates"

red= PatternFill(patternType="solid",
                fgColor="C64747")
#Crea archivo
book = Workbook()
sheet = book.active
#Carga archivo con el nombre
#book = load_workbook(FILE)
#sheet = book["Sheet"]
sheet["B1"]="Rojo"
sheet["B1"].fill=red
sheet["B1"].font = Font(color="FF0000",bold=True)
for i in range(2,13):
    sheet[f"B{i}"]=i
sheet["A3"] = "Sumatoria"    
sheet["A4"] = "=SUM(B2:B12)"
top= Side(border_style="thin")

epa= Border(top=top,right=top,bottom=top,left=top)

sheet["A4"].border =epa

#sheet3 = book[FILE_SHEET3]
#sheet2 = book[FILE_SHEET2]
sheet2 = book.create_sheet(FILE_SHEET2)
sheet3 = book.create_sheet(FILE_SHEET3)
sheet2["A1"] = "Fecha"
date = time.strftime("%x")
sheet2["A2"]=date
date = time.strftime("%y")


for i in range(1,len(array_datos)):
    date4 = time.strftime(f"%{array_datos[i]}")
    sheet3[f"A{i}"] = array_datos[i]
    sheet3[f"B{i}"] = date4



currentDir = os.getcwd()
# Open Microsoft Excel
excel = client.Dispatch("Excel.Application")
books = excel.Workbooks.Open(os.path.join(currentDir,FILE))    
ws = books.Worksheets[0]
ws.Visible = 1
ws.ExportAsFixedFormat(0,os.path.join(currentDir,"prueba.pdf") )
books.Close()
print("PPTX to PDF conversion sucessful and Saved")

# Convert into PDF File
#work_sheets.ExportAsFixedFormat(0, "ExcelTest4a.pdf")

book.save(FILE)