import string
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font 
import time

#Carga archivo con el nombre
book = load_workbook("ExcelTest3.xlsx")
#Cambiar valor
sheet = book["Hoja1"]

#Crea archivo
#book = Workbook()
# celda2 = input("Ingrese letraCelda: ")
# celda = input("Ingrese numero celda: ")
# valor = input("Ingrese valor: ")

# sheet[celda2+celda] =  valor

# #Acceder valor
# celdaobtenida = sheet[celda2+celda]
# print(celdaobtenida.value)

#Cambia tabla izq derecha
valoringresado = int(input("Ingrese valor: "))
Inicial =input("Ingrese Inicial: ")
Final =input("Ingrese Final: ")
separado_init= list(Inicial)
letra_init=separado_init[0]
letra_num=ord(letra_init.lower())-96
print(f"{letra_init}= {letra_num}")
numero_init=int(separado_init[1])

separado_fin= list(Final)
letra_fin=separado_fin[0]
letra_num2=ord(letra_fin.lower())-96
print(f"{letra_fin}= {letra_num2}")
numero_fin=int(separado_fin[1])
rango = sheet[Inicial:Final]
longitud=len(rango)
print(f"Longitud: {longitud}")
numero=0
for fila in range(numero_init,numero_fin+1):
    for col in range(letra_num,letra_num2+1):
            num_letra=chr(col+96).upper() #Celda columna
            valorobtenido= sheet[num_letra+str(fila)].value
            #print(f"{valorobtenido}")
            if(numero == valoringresado):
                print(f"{valorobtenido}")
            numero=numero+1


book.save("ExcelTest3.xlsx")