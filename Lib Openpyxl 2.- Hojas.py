from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font 
import time

#Carga archivo con el nombre
wb2 = load_workbook("ExcelTest1.xlsx")

#Crea hoja 1 
#ws1 = wb2.create_sheet("Hoja 1")
#Se crea al incio
ws2 = wb2.create_sheet("Hoja posicion 0",0)



#Atributos Hojas
#Nombre 
ws = wb2["rename"]
    #ws.title = "rename"
#Color
ws.sheet_properties.tabColor = "1072BA"

# sheet["B1"] =  a+b
# book.save")
wb2.save("ExcelTest1.xlsx")