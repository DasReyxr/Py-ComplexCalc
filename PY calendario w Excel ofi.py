from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from win32com import client
import os
import time

import win32com.client
import win32com.client.dynamic
FILE_NAME ="ExcelCal1"
EXT = ".xlsx"
DOW_CELL_INIT = "C1"
DOW_CELL_END = "I1"

book = Workbook()
sheet = book.active
#-------En caso de que ya lo haya generado comentar las dos lineas superiores
#---- Y descomentar las dos lineas inferiores
#book = load_workbook(FILE_NAME)
#sheet = book["Sheet"]

thin= Side(border_style="thin")
marco= Border(top=thin,right=thin,bottom=thin,left=thin)

sheet.column_dimensions["A"].width = 4.5
sheet.column_dimensions["B"].width = 4.5
INIT ="C2"
END ="I44"


black ="000000"
white = "FFFFFF"
gray = "808080" 
green =  "00B050"
cell_white=PatternFill(patternType="solid",
                fgColor=white)
cell_gray= PatternFill(patternType="solid",
                fgColor=gray)
cell_green= PatternFill(patternType="solid",
                fgColor=green)

for row in sheet["C1:I44"]:
    for cell in row:
        cell.border = marco

#clear sheet
for row in sheet['A1:I44']:
  for cell in row:
    cell.value = None
    cell.fill = cell_white
    cell.font = Font(color=black)

dias =[0,31,28,31,30,31,30,31,31,30,31,30,31]
dow = [0,"Lun","Mar","Mie","Jue","Vie","Sab","Dom"]
turno =["Night","Night","Free","Free","Day","Day"]
array_mes =[0,"Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]


def loop():
    mes_ingresado = int(time.strftime("%m"))
    dia_ingresado= int(time.strftime("%d"))
    print(f"Mes {mes_ingresado}) {array_mes[mes_ingresado]}")
    print(f"Dia {dia_ingresado} inicia en: ")
    dow_ingresado = int(time.strftime("%w"))
    print(f"{dow_ingresado}) {dow[dow_ingresado]} \n")
    for i in range(0,6):
        print(f"{i}) {turno[i]}")
    turno_ingresado = int(input("Ingrese turno: "))
    
    Inst_mes = Mes()
    Inst_mes.prnt_dias(mes_ingresado,dow_ingresado,turno_ingresado,dia_ingresado)
    Inst_mes.prnt_dow(dow)

    currentDir = os.getcwd()
    print(f"os {currentDir}")
    # Open Microsoft Excel
    excel = client.Dispatch("Excel.Application")
    books = excel.Workbooks.Open(os.path.join(currentDir,FILE_NAME+EXT))    
    ws = books.Worksheets[0]
    ws.Visible = 1
    DESTINY=os.path.join(currentDir,FILE_NAME+".pdf")
    ws.ExportAsFixedFormat(0,DESTINY)
    books.Close()
    print(f"PPTX to PDF {DESTINY}")
    book.save(FILE_NAME+EXT)

class Mes:
    def prnt_dow(self,array_dow):
        self.array_dow=array_dow
        Primer_letra=ord(list(INIT)[0].lower())-96
        Ult_Letra =ord(list(END)[0].lower())-96
        for char in range(Primer_letra,Ult_Letra+1):
                num_letra=chr(char+96).upper() #Celda columna  
                sheet[f"{num_letra}1"] =dow[char-2]
                      
    def prnt_dias(self,mes,dow_init,turn_init,num_init):
        vuelta = False
        self.mes = mes
        self.dow_init = dow_init  
        self.turn_init = turn_init    
        numero_init=int(list(INIT)[1])
        numero_fin=int(list(END)[1]+list(END)[2])
        letra_fin = ord(list(END)[0].lower())-96
        for fila in range(numero_init,numero_fin+1):
            if (vuelta==False):
                letra_init=ord(list(INIT)[0].lower())-96+dow_init-1
            else:
                letra_init = ord(list(INIT)[0].lower())-96
            for col in range(letra_init,letra_fin+1):
                num_letra=chr(col+96).upper() #Celda columna
                if(self.dow_init>7):
                    self.dow_init=1
                if(self.turn_init==6):
                    self.turn_init=0
                
                if(num_init>(dias[self.mes])):
                    cc1=f"A"+str(fila-3)
                    cc2=f"B"+str(fila)
                    sheet[cc1] = array_mes[self.mes]
                    sheet[cc1].border = marco
                    sheet.merge_cells(f"{cc1}:{cc2}")
                    sheet[cc1].alignment = Alignment(horizontal="center",vertical="center")
                    if(self.mes==12):
                        self.mes=1
                        num_init=1    
                    else:
                        self.mes=self.mes+1
                        num_init=1
                cell= f"{num_letra}"+str(fila)
                sheet.column_dimensions[num_letra].width = 7
                sheet[cell] = num_init
                if (self.turn_init==0):
                    sheet[cell].fill = cell_gray
                    sheet[cell].font = Font(color=white)           
                elif (self.turn_init==1):        
                    sheet[cell].fill = cell_gray
                    sheet[cell].font = Font(color=white)
                elif (self.turn_init==2):
                    sheet[cell].fill = cell_green
                    sheet[cell].font = Font(color=white)
                elif (self.turn_init==3):
                    sheet[cell].fill = cell_green
                    sheet[cell].font = Font(color=white)
                elif (self.turn_init==4):
                    sheet[cell].fill = cell_white
                    sheet[cell].font = Font(color=black)
                elif (self.turn_init==5):
                    sheet[cell].fill = cell_white
                    sheet[cell].font = Font(color=black)
                self.dow_init=self.dow_init+1
                self.turn_init=self.turn_init+1
                num_init=num_init+1
            vuelta = True

loop()